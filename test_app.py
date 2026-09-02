import os
import openpyxl
from scraper import extract_contact_info, PROVINSI_LIST, fetch_kabupaten
from export import export_to_excel, export_to_csv, clean_whatsapp

def test_whatsapp_cleaning():
    print("Testing WhatsApp cleaning...")
    wa1, link1 = clean_whatsapp("0812-3456-7890")
    assert wa1 == "6281234567890", f"Failed: {wa1}"
    assert link1 == "https://wa.me/6281234567890", f"Failed: {link1}"

    wa2, link2 = clean_whatsapp("+62 857 1122 3344")
    assert wa2 == "6285711223344", f"Failed: {wa2}"
    assert link2 == "https://wa.me/6285711223344", f"Failed: {link2}"
    print("WhatsApp cleaning PASSED!")

def test_extract_contact_info():
    print("Testing contact info regex extraction...")
    sample_html = """
    <div class="card">
        <h3>PT MAJU JAYA KONSTRUKSI</h3>
        <table>
            <tr><th>Alamat</th><td>Jl. Gatot Subroto No. 45, Jakarta Selatan</td></tr>
            <tr><th>Penanggung Jawab</th><td>Ir. Budi Santoso</td></tr>
            <tr><th>Kontak</th><td>Hubungi kami di WA: 0812-8899-7766 atau Telp (021) 5251234</td></tr>
            <tr><th>Email</th><td>info@majujaya.co.id, sekretariat@majujaya.co.id</td></tr>
        </table>
    </div>
    """
    res = extract_contact_info(sample_html)
    print("Extracted:", res)
    assert "6281288997766" in res["whatsapp"], "WA extraction failed"
    assert "https://wa.me/6281288997766" in res["wa_link"], "WA link failed"
    assert "info@majujaya.co.id" in res["email"], "Email extraction failed"
    assert "sekretariat@majujaya.co.id" in res["email"], "Second email extraction failed"
    assert "(021) 5251234" in res["telepon"], "Phone extraction failed"
    assert "Jl. Gatot Subroto" in res["alamat"], "Alamat failed"
    assert "Budi Santoso" in res["pimpinan"], "Pimpinan failed"
    print("Contact info extraction PASSED!")

def test_excel_export():
    print("Testing Excel and CSV export...")
    data = [
        {
            "no": 1,
            "nama": "PT WIJAYA KARYA (PERSERO) TBK",
            "whatsapp": "6281234567890",
            "wa_link": "https://wa.me/6281234567890",
            "email": "corporate@wika.co.id",
            "telepon": "(021) 8192808",
            "pimpinan": "Agung Budi",
            "provinsi": "DKI Jakarta",
            "kabupaten": "Jakarta Timur",
            "npwp": "01.001.614.5-093.000",
            "kualifikasi": "BESAR",
            "alamat": "Jl. D.I. Panjaitan Kav. 9-10",
            "subklas": "BG001, SI001"
        }
    ]
    os.makedirs("test_out", exist_ok=True)
    excel_path = "test_out/test_export.xlsx"
    csv_path = "test_out/test_export.csv"

    export_to_excel(data, excel_path)
    assert os.path.exists(excel_path), "Excel file was not created"
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    assert ws.cell(row=2, column=2).value == "PT WIJAYA KARYA (PERSERO) TBK"
    assert ws.cell(row=2, column=4).hyperlink is not None, "Hyperlink not set"
    print("Excel export PASSED!")

    export_to_csv(data, csv_path)
    assert os.path.exists(csv_path), "CSV file was not created"
    print("CSV export PASSED!")

if __name__ == "__main__":
    test_whatsapp_cleaning()
    test_extract_contact_info()
    test_excel_export()
    print("\nALL VERIFICATION TESTS PASSED SUCCESSFULLY!")
