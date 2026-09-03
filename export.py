import os
import re
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_HEADERS = [
    "No", "Nama Badan Usaha", "WhatsApp (62...)", "Link WhatsApp",
    "Email Perusahaan", "No Telepon Kantor", "Pimpinan / PJBU",
    "Provinsi", "Kabupaten / Kota", "NPWP", "Kualifikasi",
    "Alamat Lengkap", "Status / Subklasifikasi"
]

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
REGULAR_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER_COLS = {1, 3, 4, 8, 9, 10, 11}


def clean_whatsapp(number_str):
    if not number_str:
        return "", ""
    digits = re.sub(r"[^\d]", "", str(number_str))
    if digits.startswith("08"):
        digits = "628" + digits[2:]
    elif digits.startswith("8") and len(digits) >= 9:
        digits = "62" + digits
    elif digits.startswith("0") and len(digits) >= 9:
        digits = "62" + digits[1:]
    if digits.startswith("628") and len(digits) >= 10:
        return digits, f"https://wa.me/{digits}"
    return digits, ""


def export_to_excel(data_list, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Kontak LPJK"
    ws.append(EXCEL_HEADERS)

    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, item in enumerate(data_list, start=2):
        wa_num = item.get("whatsapp", "")
        wa_link = item.get("wa_link", "")
        if not wa_link and wa_num:
            _, wa_link = clean_whatsapp(wa_num)

        ws.append([
            item.get("no", row_idx - 1),
            item.get("nama", ""),
            wa_num,
            "Chat WA" if wa_link else "-",
            item.get("email", ""),
            item.get("telepon", ""),
            item.get("pimpinan", ""),
            item.get("provinsi", ""),
            item.get("kabupaten", ""),
            item.get("npwp", ""),
            item.get("kualifikasi", ""),
            item.get("alamat", ""),
            item.get("subklas", ""),
        ])

        for col_idx in range(1, len(EXCEL_HEADERS) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = REGULAR_FONT
            c.border = THIN_BORDER
            align = Alignment(horizontal="center" if col_idx in CENTER_COLS else "left", vertical="center")
            c.alignment = align

            if col_idx == 4 and wa_link:
                c.hyperlink = wa_link
                c.font = LINK_FONT

            if col_idx == 5 and item.get("email"):
                c.hyperlink = f"mailto:{item.get('email')}"
                c.font = LINK_FONT

    ws.row_dimensions[1].height = 28
    for r in range(2, len(data_list) + 2):
        ws.row_dimensions[r].height = 20

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)

    wb.save(output_path)
    return output_path


def export_to_csv(data_list, output_path):
    pd.DataFrame(data_list).to_csv(output_path, index=False, encoding="utf-8-sig")
    return output_path


def export_to_json(data_list, output_path):
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data_list, f, ensure_ascii=False, indent=2)
    return output_path
